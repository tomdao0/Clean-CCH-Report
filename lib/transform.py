import pandas as pd
import urllib.parse
import sqlalchemy
from sqlalchemy import create_engine, text
import openpyxl
import re
from datetime import datetime
import pytz
import os
import polars as pl


class ARBalanceListing:

    def __init__(self, folder_path):
        self.folder_path = folder_path

    def clean_column_name(self, col_name):
        return re.sub(r"[^a-zA-Z0-9]", "", col_name)

    def process_files(self):
        xlsx_files = [
            f
            for f in os.listdir(self.folder_path)
            if f.endswith(".xlsx") and not (f.startswith("~"))
        ]
        df_list = []
        # Process each file
        for file in xlsx_files:
            file_path = os.path.join(self.folder_path, file)
            temp = pl.read_excel(file_path, sheet_id=2)
            # df = pd.read_excel(file_path, sheet_name=1)
            df = temp.to_pandas()
            df = df.astype(str)
            # Find row contains header
            found = False
            for row in df.index:
                for col_idx in range(df.shape[1]):
                    if df.iloc[row, col_idx]:
                        if df.iloc[row, col_idx] == "Client ID":
                            row_header = row
                            found = True
                            break
                if found:
                    break
            # List all cols name to keep
            keep_cols = []
            for col_idx in range(df.shape[1]):
                if df.iloc[row_header, col_idx] != "None":
                    keep_cols.append(col_idx)
            found = False
            df = df.iloc[:, keep_cols]

            df.columns = [
                self.clean_column_name(df.iloc[row_header, col])
                for col in range(df.shape[1])
            ]
            df = df[
                ((df["ClientID"] != "None") | (df["TransactionDate"] != "None"))
                & (df["ClientID"] != "Client ID")
                & (df["ClientID"] != "Grand totals:")
            ].reset_index(drop=True)
            df = df.replace(["None", ""], None)
            df["ClientID"] = df["ClientID"].ffill()
            df["ClientID"] = df["ClientID"].apply(
                lambda x: (
                    x.split()[4] if isinstance(x, str) and len(x.split()) > 4 else None
                )
            )
            df = df[df["TransactionDate"].notna()]
            df.drop(
                columns=["ClientName", "ARBalance", "AccountingPeriodDate"],
                inplace=True,
            )
            df.rename(
                columns={
                    "ClientID": "ClientIdSubId",
                    "Document": "TransNumber",
                    "AppliedTo": "AppliedNumber",
                },
                inplace=True,
            )
            df["TransactionDate"] = pd.to_datetime(df["TransactionDate"])
            df_list.append(df)
        final_df = pd.concat(df_list, ignore_index=True)
        final_df["Amount"] = pd.to_numeric(final_df["Amount"])
        return final_df


class StaffPosted:

    def __init__(self, folder_path):
        self.folder_path = folder_path

    def substring_after_5th_whitespace(self, txt):
        parts = txt.split(" ", 3)
        if len(parts) > 3:
            return parts[3]
        return None

    def clean_column_name(self, col_name):
        return re.sub(r"[^a-zA-Z0-9]", "", col_name)

    def process_files(self):
        # List to store DataFrames
        df_list = []
        # Get list of .xlsx files
        xlsx_files = [
            f
            for f in os.listdir(self.folder_path)
            if f.endswith(".xlsx") and not (f.startswith("~"))
        ]

        # Process each file

        for file in xlsx_files:
            file_path = os.path.join(self.folder_path, file)
            df = pd.read_excel(file_path, sheet_name=1)
            df = df.astype(str)
            found = False
            for row in range(16):  # Because this cell always before 16 rows
                for col in df.columns:
                    if df.at[row, col] and isinstance(df.at[row, col], str):
                        if df.at[row, col][:28] == "For Accounting period dates:":
                            match = re.search(
                                r"For Transaction dates:(\d{1,2}/\d{1,2}/\d{4}) - (\d{1,2}/\d{1,2}/\d{4})",
                                df.at[row, col],
                            )
                            if match:
                                begin_date = match.group(1)
                                end_date = match.group(2)
                                found = True
                                break
                            else:
                                print("Transaction dates not found.")
                if found:
                    break
            # Find Header row
            found = False
            for row in df.index:
                for col_idx in range(df.shape[1]):
                    if df.iloc[row, col_idx]:
                        if df.iloc[row, col_idx] == "Hours":
                            row_header = row
                            found = True
                            break
                if found:
                    break

            # List all cols name to keep
            keep_cols = []
            for col_idx in range(df.shape[1]):
                if df.iloc[row_header, col_idx] != "nan":
                    keep_cols.append(col_idx)
            df = df.iloc[:, keep_cols]
            # Exclude rows has "Grand Total" to the end
            for row in reversed(df.index):
                if df.iloc[row, 0] == "Grand Totals:":
                    df.drop(df.index[row:], inplace=True)
                    break
            df.columns = [
                self.clean_column_name(df.iloc[row_header, col])
                for col in range(df.shape[1])
            ]
            df.drop(df.index[: row_header + 1], inplace=True)
            df["StaffID"] = df["PostedHours"]
            df["BankedUsedHours"] = df["BankedUsedHours"].apply(
                lambda x: x.split(" ")[1] if len(x.split(" ")) > 1 else x
            )
            df["TypeBankedHrs"] = df["BankedUsedHours"].apply(
                lambda x: x.split(" ")[0] if len(x.split(" ")) > 1 else None
            )
            df = df[df["StaffID"].str[-1:] != ")"]
            df["StaffID"] = df["StaffID"].apply(
                lambda x: self.substring_after_5th_whitespace(x)
            )
            df["StaffID"] = df["StaffID"].ffill()
            df = df[df["Hours"] != "nan"].reset_index(drop=True)
            df.drop(columns="", inplace=True)
            df.rename(
                columns={"BankedUsedHours": "BankedHoursUsed", "Hours": "BillHours"},
                inplace=True,
            )
            df["begin_date"] = datetime.strptime(begin_date, "%m/%d/%Y")
            df["end_date"] = datetime.strptime(end_date, "%m/%d/%Y")
            df_list.append(df)
        final_df = pd.concat(df_list, ignore_index=True)
        for i in final_df.columns:
            if i not in ["StaffID", "begin_date", "end_date"]:
                final_df[i] = pd.to_numeric(final_df[i], errors="coerce")
        return final_df


class WIPActivity:
    def __init__(self, folder_path):
        self.folder_path = folder_path

    def find_date(self, txt):
        dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", txt)
        first_date = dates[1] if dates else None
        return first_date

    def clean_column_name(self, col_name):
        return re.sub(r"[^a-zA-Z0-9]", "", col_name)

    def process_files(self):
        # List to store DataFrames
        df_list = []

        # Get list of .xlsx files
        xlsx_files = [
            f
            for f in os.listdir(self.folder_path)
            if f.endswith(".xlsx") and not (f.startswith("~"))
        ]
        # From here to process each file: This code will be determine what position of special cell in this firm
        file_path = os.path.join(self.folder_path, xlsx_files[0])
        temp = pl.read_excel(file_path, sheet_id=2)
        df = temp.to_pandas()
        # df = pd.read_excel(file_path, sheet_name=1)
        df = df.astype(str)

        # Find row contains header
        found = False
        for row in df.index:
            for col_idx in range(df.shape[1]):
                if df.iloc[row, col_idx]:
                    if df.iloc[row, col_idx] == "WIP Beg Balance":
                        row_header = row
                        found = True
                        break
            if found:
                break
        # List all cols name to keep
        # Keep ClientIdSubId column
        keep_cols = []
        found = False
        for row in range(row_header, df.shape[0]):
            for col_idx in range(df.shape[1]):
                if df.iloc[row, col_idx]:
                    if df.iloc[row, col_idx][:18] == "Client ID Sub ID :":
                        keep_cols.append(col_idx)
                        found = True
                        break
            if found:
                break
        for col_idx in range(df.shape[1]):
            if (df.iloc[row_header, col_idx] != "None") and (
                df.iloc[row_header, col_idx] != ""
            ):
                keep_cols.append(col_idx)
        # Determine CutOffDate
        found = False
        for row in range(16):  # Because this cell always before 16 rows
            for col in df.columns:
                if df.at[row, col] and isinstance(df.at[row, col], str):
                    if df.at[row, col][:3] == "PTD":

                        row_cutoff, col_cutoff = row, col
                        found = True
                        break
            if found:
                break

        # Process each file
        for file in xlsx_files:
            file_path = os.path.join(self.folder_path, file)
            temp = pl.read_excel(file_path, sheet_id=2)
            df = temp.to_pandas()

            # df = pd.read_excel(file_path, sheet_name=1)
            CutOffDate = self.find_date(df.at[row_cutoff, col_cutoff])
            df = df.astype(str)
            df = df.iloc[:, keep_cols]
            df.columns = [
                self.clean_column_name(df.iloc[row_header, col])
                for col in range(df.shape[1])
            ]

            df.columns.values[0] = "Type"
            df["ClientIdSubId"] = df["Type"]

            df = df[
                (df["Type"] != "None") & (df["Type"] != "RTD") & (df["Type"] != "")
            ].reset_index(drop=True)

            # Exclude rows has "Grand Total" to the end
            for row in reversed(df.index):
                if df.iloc[row, 0] == "Grand Totals":
                    df.drop(df.index[row:], inplace=True)
                    break

            df["ClientIdSubId"] = df["ClientIdSubId"].replace(["PTD", ""], None)
            df["ClientIdSubId"] = df["ClientIdSubId"].ffill()
            df["ClientIdSubId"] = df["ClientIdSubId"].apply(lambda x: x.split(" ")[5])
            df["CutOffDate"] = datetime.strptime(CutOffDate, "%m/%d/%Y")
            df = df[df["Type"] == "PTD"]
            df = df.drop(columns=["Type", "WIP", "RelievedWIPAdjust"]).reset_index(
                drop=True
            )
            df_list.append(df)
        final_df = pd.concat(df_list, ignore_index=True)
        for i in final_df.columns:
            if i not in ["ClientIdSubId", "CutOffDate"]:
                final_df[i] = pd.to_numeric(final_df[i], errors="coerce")
        return final_df


class StaffList:

    def __init__(self, file_path):
        self.file_path = file_path

    def process_files(self):
        # Process each Staff
        file_path = self.file_path
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.worksheets[1]
        last_row = sheet.max_row
        D = {}

        # Find last row with "Pay Type:"
        for row in range(last_row - 6, last_row + 1):
            cell = sheet[f"C{row}"]
            if cell.value == "Pay Type:":
                last_row = row + 1
                break

        # Find first row with "StaffID"
        for row in range(1, last_row):
            if sheet[f"C{row}"].value is not None:
                if sheet[f"C{row}"].value == "Staff ID":
                    first_row = row + 1  # First position of StaffID
                    break
        # Find first row with "Full Name:"
        first_range = 0
        last_range = 0
        for row in range(1, last_row):
            if sheet[f"C{row}"].value is not None:
                if sheet[f"C{row}"].value == "Full Name:":
                    if first_range == 0 and last_range == 0:
                        first_range = row
                    elif first_range != 0 and last_range == 0:
                        last_range = row
                    else:
                        size_range = last_range - first_range
                        break
        # Process the data
        for i in range(1, int((last_row - first_row + 1) / size_range + 1)):
            StaffID = sheet[f"C{(i-1)*size_range + first_row + 1}"].value
            ReportName = sheet[f"D{(i-1)*size_range + first_row + 1}"].value
            StaffNameNull = sheet[f"D{(i-1)*size_range + first_row + 2}"].value
            StaffOffice = sheet[f"D{(i-1)*size_range + first_row + 4}"].value
            StaffBU = sheet[f"D{(i-1)*size_range + first_row + 5}"].value
            StaffDepartment = sheet[f"D{(i-1)*size_range + first_row + 6}"].value
            ReportingManager = sheet[f"K{(i-1)*size_range + first_row + 5}"].value
            StaffStatus = sheet[f"O{(i-1)*size_range + first_row + 1}"].value
            D[i] = [
                StaffID,
                ReportName,
                StaffNameNull,
                StaffOffice,
                StaffBU,
                StaffDepartment,
                ReportingManager,
                StaffStatus,
            ]

        # Create DataFrame
        df = pd.DataFrame.from_dict(
            D,
            orient="index",
            columns=[
                "StaffID",
                "ReportName",
                "StaffNameNull",
                "StaffOffice",
                "StaffBU",
                "StaffDepartment",
                "ReportingManager",
                "StaffStatus",
            ],
        )
        return df


class StaffMonthly:

    def __init__(self, folder_path, utcFormat):
        self.folder_path = folder_path
        self.utcFormat = utcFormat

    def process_files(self):
        # Define timezone and current time based on UTC
        utc_minus = pytz.timezone(self.utcFormat)
        current_time_utc_minus = datetime.now(utc_minus).strftime("%Y-%m-%d %H:%M:%S")

        # List to store DataFrames
        df_list = []

        # Get list of .xlsx files
        xlsx_files = [
            f
            for f in os.listdir(self.folder_path)
            if f.endswith(".xlsx") and not (f.startswith("~"))
        ]

        # Process each file
        for file in xlsx_files:
            file_path = os.path.join(self.folder_path, file)
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.worksheets[1]
            last_row = sheet.max_row
            D = {}
            l = [
                "Production Hours",
                "Production Amounts",
                "Billed Hours",
                "Billed Amounts",
                "Billed Write +/- Amounts",
            ]
            # Find last row with "Grand Totals :"
            for row in range(last_row - 12, last_row + 1):
                cell = sheet[f"B{row}"]
                if cell.value == "Grand Totals":
                    last_row = row - 1
                    break

            # Find first row with "Staff ID"
            for row in range(1, last_row):
                if sheet[f"B{row}"].value is not None:
                    if sheet[f"B{row}"].value[:8] == "Staff ID":
                        first_row = row
                        break

            # Find first row with value starting with "For the Dates"
            for row in range(1, 20):
                if sheet[f"H{row}"].value is not None:
                    if sheet[f"H{row}"].value[:13] == "For the Dates":
                        text = sheet[f"H{row}"].value
                        dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", text)
                        first_date = dates[0] if dates else None
                        break

            # Process the data
            for i in range(1, int((last_row - first_row) / 6 + 1)):
                # Because first staff contain year of report so we exclude first record
                if i == 1:
                    StaffID = sheet[f"B{first_row}"].value.split(" ")[3]
                    for j in range(5):
                        Type = l[j]
                        Jan = sheet[f"D{first_row + 2 + j}"].value
                        Feb = sheet[f"E{first_row + 2 + j}"].value
                        Mar = sheet[f"F{first_row + 2 + j}"].value
                        Apr = sheet[f"I{first_row + 2 + j}"].value
                        May = sheet[f"J{first_row + 2 + j}"].value
                        Jun = sheet[f"K{first_row + 2 + j}"].value
                        Jul = sheet[f"L{first_row + 2 + j}"].value
                        Aug = sheet[f"M{first_row + 2 + j}"].value
                        Sep = sheet[f"N{first_row + 2 + j}"].value
                        Oct = sheet[f"Q{first_row + 2 + j}"].value
                        Nov = sheet[f"R{first_row + 2 + j}"].value
                        Dec = sheet[f"S{first_row + 2 + j}"].value
                        Total = sheet[f"T{first_row + 2 + j}"].value
                        D[j + 1] = [
                            StaffID,
                            Type,
                            Jan,
                            Feb,
                            Mar,
                            Apr,
                            May,
                            Jun,
                            Jul,
                            Aug,
                            Sep,
                            Oct,
                            Nov,
                            Dec,
                            Total,
                        ]
                else:
                    StaffID = sheet[f"B{first_row + (i-1) * 6 + 1}"].value.split(" ")[3]
                    for j in range(5):
                        Type = l[j]
                        Jan = sheet[f"D{first_row + (i-1) * 6 + 2 + j}"].value
                        Feb = sheet[f"E{first_row + (i-1) * 6 + 2 + j}"].value
                        Mar = sheet[f"F{first_row + (i-1) * 6 + 2 + j}"].value
                        Apr = sheet[f"I{first_row + (i-1) * 6 + 2 + j}"].value
                        May = sheet[f"J{first_row + (i-1) * 6 + 2 + j}"].value
                        Jun = sheet[f"K{first_row + (i-1) * 6 + 2 + j}"].value
                        Jul = sheet[f"L{first_row + (i-1) * 6 + 2 + j}"].value
                        Aug = sheet[f"M{first_row + (i-1) * 6 + 2 + j}"].value
                        Sep = sheet[f"N{first_row + (i-1) * 6 + 2 + j}"].value
                        Oct = sheet[f"Q{first_row + (i-1) * 6 + 2 + j}"].value
                        Nov = sheet[f"R{first_row + (i-1) * 6 + 2 + j}"].value
                        Dec = sheet[f"S{first_row + (i-1) * 6 + 2 + j}"].value
                        Total = sheet[f"T{first_row + (i-1) * 6 + 2 + j}"].value
                        D[(i - 1) * 6 + j + 1] = [
                            StaffID,
                            Type,
                            Jan,
                            Feb,
                            Mar,
                            Apr,
                            May,
                            Jun,
                            Jul,
                            Aug,
                            Sep,
                            Oct,
                            Nov,
                            Dec,
                            Total,
                        ]
            # Create DataFrame
            df = pd.DataFrame.from_dict(
                D,
                orient="index",
                columns=[
                    "StaffID",
                    "Type",
                    "Jan",
                    "Feb",
                    "Mar",
                    "Apr",
                    "May",
                    "Jun",
                    "Jul",
                    "Aug",
                    "Sep",
                    "Oct",
                    "Nov",
                    "Dec",
                    "Total",
                ],
            )
            df.fillna(0.0, inplace=True)
            df["RunningTime"] = current_time_utc_minus
            df["CutOff"] = datetime.strptime(first_date, "%m/%d/%Y")

            df_list.append(df)

        final_df = pd.concat(df_list, ignore_index=True)
        for i in final_df.columns:
            if not (i in ["StaffID", "Type", "RunningTime", "CutOff"]):
                final_df[i] = (
                    final_df[i]
                    .astype(str)
                    .str.replace(",", "", regex=False)
                    .astype(float)
                )
            if i in ["RunningTime", "CutOff"]:
                final_df[i] = pd.to_datetime(final_df[i])
        return final_df


class WIPARRecon:

    def __init__(self, folder_path):
        self.folder_path = folder_path

    def find_date(self, txt):
        dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", txt)
        first_date = dates[1] if dates else None
        return first_date

    def clean_column_name(self, col_name):
        return re.sub(r"[^a-zA-Z]", "", col_name)

    def process_files(self):
        # List to store DataFrames
        df_list = []

        # Get list of .xlsx files
        xlsx_files = [
            f
            for f in os.listdir(self.folder_path)
            if f.endswith(".xlsx") and not (f.startswith("~"))
        ]
        # From here to process each file: This code will be determine what position of special cell in this firm
        file_path = os.path.join(self.folder_path, xlsx_files[0])
        # df = pd.read_excel(file_path, sheet_name=1)
        temp = pl.read_excel(file_path, sheet_id=2)
        df = temp.to_pandas()
        df = df.astype(str)

        # Find row contains header
        found = False
        for row in df.index:
            for col_idx in range(df.shape[1]):
                if df.iloc[row, col_idx]:
                    if df.iloc[row, col_idx] == "WIP Beg\nBalance":
                        row_header = row
                        found = True
                        break
            if found:
                break
        # Determine cols name to keep
        keep_cols = []
        for col_idx in range(df.shape[1]):
            if (df.iloc[row_header, col_idx] != "None") and (
                df.iloc[row_header, col_idx] != ""
            ):
                keep_cols.append(col_idx)
        # Determine CutOffDate
        found = False

        for row in range(16):  # Because this cell always before 16 rows
            for col in df.columns:
                if df.at[row, col] and isinstance(df.at[row, col], str):
                    if df.at[row, col][:28] == "For Accounting period dates:":
                        row_cutoff, col_cutoff = row, col
                        found = True
                        break  # Break the inner loop
            if found:
                break
        # Process each file
        for file in xlsx_files:
            file_path = os.path.join(self.folder_path, file)
            temp = pl.read_excel(file_path, sheet_id=2)
            df = temp.to_pandas()
            # df = pd.read_excel(file_path, sheet_name=1)
            CutOffDate = self.find_date(df.at[row_cutoff, col_cutoff])
            df = df.astype(str)
            df = df.iloc[:, keep_cols]
            df.columns = [
                self.clean_column_name(df.iloc[row_header, col])
                for col in range(df.shape[1])
            ]
            df = df.drop(df.index[: row_header + 1]).reset_index(drop=True)
            df["ClientIdSubId"] = df["WIPBegBalance"]
            # Exclude rows has "Grand Total" to the end
            for row in reversed(df.index):
                if df.iloc[row, 0] == "Grand Totals":
                    df.drop(df.index[row:], inplace=True)
                    break
            for row in df.index:
                if len(df.at[row, "ClientIdSubId"].split(" ")) == 6:
                    df.at[row, "ClientIdSubId"] = df.at[row, "ClientIdSubId"].split(
                        " "
                    )[5]
                else:
                    df.at[row, "ClientIdSubId"] = df.at[row - 1, "ClientIdSubId"]
            df = df[(df["Hours"] != "None") & (df["Hours"] != "")]
            df["RealPercent"] = df["RealPercent"].replace("%", "", regex=True)
            df["CutOff"] = datetime.strptime(CutOffDate, "%m/%d/%Y")
            df_list.append(df)
        final_df = pd.concat(df_list, ignore_index=True)

        for i in final_df.columns:
            if i not in ["CutOff", "ClientIdSubId"]:
                final_df[i] = pd.to_numeric(final_df[i], errors="coerce")
        df["CutOff"] = pd.to_datetime(df["CutOff"])
        final_df.rename(
            columns={
                "WIPBegBalance": "WIPBegin",
                "WriteUpWriteDown": "WriteUD",
                "WIPEndBalance": "WIPEnd",
                "ARBegBalance": "ARBegin",
                "InvoicewSalesTax": "InvoiceSalesTax",
                "Adjustments": "Adjustment",
                "FinanceCharges": "Charges",
                "AREndBalance": "AREnd",
            },
            inplace=True,
        )
        return final_df


class WIPARAging:

    def __init__(self, folder_path):
        self.folder_path = folder_path

    def find_date(self, txt):
        dates = re.findall(r"\d{1,2}/\d{1,2}/\d{4}", txt)
        first_date = dates[1] if dates else None
        return first_date

    def clean_column_name(self, col_name):
        return re.sub(r"[^a-zA-Z0-9]", "", col_name)

    def get_payment(self, txt):
        date_match = re.search(r"\d{1,2}/\d{1,2}/\d{4}", txt)
        amount_match = re.search(r"\$\d{1,3}(?:,\d{3})*(?:\.\d{2})?", txt)
        LastPaymentDate = date = date_match.group() if date_match else None
        LastPaymentAmount = (
            amount_match.group().replace("$", "").replace(",", "")
            if amount_match
            else None
        )
        return LastPaymentDate, LastPaymentAmount

    def process_files(self):
        # List to store DataFrames
        df_list = []

        # Get list of .xlsx files
        xlsx_files = [
            f
            for f in os.listdir(self.folder_path)
            if f.endswith(".xlsx") and not (f.startswith("~"))
        ]
        # From here to process each file: This code will be determine what position of special cell in this firm
        file_path = os.path.join(self.folder_path, xlsx_files[0])
        temp = pl.read_excel(file_path, sheet_id=2)
        # df = pd.read_excel(file_path, sheet_name=1)
        df = temp.to_pandas()
        df = df.astype(str)

        # Find row contains header
        found = False
        for row in df.index:
            for col_idx in range(df.shape[1]):
                if df.iloc[row, col_idx]:
                    if df.iloc[row, col_idx] == "Total":
                        row_header = row
                        found = True
                        break
            if found:
                break
        # List all cols name to keep
        # Keep ClientIdSubId column
        keep_cols = []
        found = False
        for row in range(row_header, df.shape[0]):
            for col_idx in range(df.shape[1]):
                if df.iloc[row, col_idx]:
                    if df.iloc[row, col_idx][:18] == "Client ID Sub ID :":
                        keep_cols.append(col_idx)
                        found = True
                        break
            if found:
                break
        # Keep type of transaction column
        found = False
        for row in range(row_header, df.shape[0]):
            for col_idx in range(df.shape[1]):
                if df.iloc[row, col_idx]:
                    if df.iloc[row, col_idx][:18] == "AR":
                        keep_cols.append(col_idx)
                        found = True
                        break
            if found:
                break
        # Keep all column has title
        for col_idx in range(df.shape[1]):
            if df.iloc[row_header, col_idx] != "nan":
                keep_cols.append(col_idx)

        # Determine CutOffDate
        found = False
        for row in range(16):  # Because this cell always before 16 rows
            for col in df.columns:
                if df.at[row, col] and isinstance(df.at[row, col], str):
                    if df.at[row, col][:20] == "For WIP dates as of:":

                        row_cutoff, col_cutoff = row, col
                        found = True
                        break  # Break the inner loop
            if found:
                break

        # Process each file
        for file in xlsx_files:
            file_path = os.path.join(self.folder_path, file)
            temp = pl.read_excel(file_path, sheet_id=2)
            df = temp.to_pandas()
            # df = pd.read_excel(file_path, sheet_name=1)
            CutOffDate = self.find_date(df.at[row_cutoff, col_cutoff])
            df = df.astype(str)
            df = df.iloc[:, keep_cols]
            df.columns = [
                self.clean_column_name(df.iloc[row_header, col])
                for col in range(df.shape[1])
            ]
            df.columns.values[0] = "LastPaymentDate"
            df.columns.values[1] = "Type"
            df["ClientIdSubId"] = df["LastPaymentDate"]
            df["LastPaymentAmount"] = df["LastPaymentDate"]
            df = df.drop(df.index[: row_header + 1]).reset_index(drop=True)
            # Exclude rows has "Grand Total" to the end
            for row in reversed(df.index):
                if df.iloc[row, 0] == "Grand Totals :":
                    df.drop(df.index[row:], inplace=True)
                    break
            df = df[(df["LastPaymentDate"] != "nan") | (df["Type"] != "nan")]
            df["ClientIdSubId"] = (
                df["ClientIdSubId"]
                .str.split()
                .apply(lambda x: x[5] if len(x) > 5 else None)
            )
            df["LastPaymentAmount"] = df["LastPaymentAmount"].apply(
                lambda x: self.get_payment(x)[1]
            )
            df["LastPaymentDate"] = df["LastPaymentDate"].apply(
                lambda x: self.get_payment(x)[0]
            )
            df["ClientIdSubId"] = df["ClientIdSubId"].ffill()
            df["LastPaymentAmount"] = df["LastPaymentAmount"].ffill()
            df["LastPaymentDate"] = df["LastPaymentDate"].ffill()
            df = df[(df["Type"] == "WIP") | (df["Type"] == "AR")]
            df.drop(columns=["", "None"], inplace=True)
            df = df.pivot(index="ClientIdSubId", columns="Type")
            df = df.reset_index()
            df.columns = [
                f"{col[0]}_{col[1]}" if col[1] != "" else col[0] for col in df.columns
            ]
            df.drop(
                columns=["LastPaymentDate_AR", "LastPaymentAmount_AR"], inplace=True
            )
            df.rename(
                columns={
                    "LastPaymentAmount_WIP": "LastPaymentAmount",
                    "LastPaymentDate_WIP": "LastPaymentDate",
                },
                inplace=True,
            )
            df["LastPaymentDate"] = pd.to_datetime(
                df["LastPaymentDate"], format="%m/%d/%Y"
            )
            df["CutOffDate"] = datetime.strptime(CutOffDate, "%m/%d/%Y")
            df_list.append(df)
        final_df = pd.concat(df_list, ignore_index=True)
        for i in final_df.columns:
            if i not in [
                "ClientIdSubId",
                "LastPaymentAmount",
                "LastPaymentDate",
                "CutOffDate",
            ]:
                final_df[i] = pd.to_numeric(final_df[i], errors="coerce")
        final_df = final_df.rename(
            columns={
                "Total_AR": "ARTotal",
                "Current030_AR": "AR0030",
                "2ndAging3160_AR": "AR3160",
                "3rdAging6190_AR": "AR6190",
                "4thAging91120_AR": "AR91120",
                "5thAging121150_AR": "AR121150",
                "6thAging151180_AR": "AR151180",
                "7thAgingOver181_AR": "AROver180",
                "Total_WIP": "WIPTotal",
                "Current030_WIP": "WIP0030",
                "2ndAging3160_WIP": "WIP3160",
                "3rdAging6190_WIP": "WIP6190",
                "4thAging91120_WIP": "WIP91120",
                "5thAging121150_WIP": "WIP121150",
                "6thAging151180_WIP": "WIP151180",
                "7thAgingOver181_WIP": "WIPOver180",
                "CutOffDate": "CutOff",
            }
        )
        return final_df


class CreateTableInSQLServer:
    def __init__(self, SQLServerName, DBName, TableName, UserName, PWD, df_data):
        self.connect_string = urllib.parse.quote_plus(
            "DRIVER={ODBC Driver 17 for SQL Server};"
            "Server=" + SQLServerName + ";"
            "Database=" + DBName + ";"
            "UID=" + UserName + ";"
            "PWD=" + PWD + ";"
        )
        self.TableName = TableName
        self.df_data = df_data

    def run(self):
        engine = create_engine(
            f"mssql+pyodbc:///?odbc_connect={self.connect_string}",
            fast_executemany=True,
        )
        with engine.connect() as connection:
            self.df_data.to_sql(
                self.TableName, connection, index=False, if_exists="replace"
            )
            print("OK")
